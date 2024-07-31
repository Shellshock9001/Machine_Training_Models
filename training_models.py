import json
import numpy as np
import os
import csv
import joblib
import time
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.linear_model import SGDClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, roc_curve, auc, classification_report, confusion_matrix
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

class IncrementalTrainer:
    """
    Class to handle incremental training of models using partial fitting.
    This class loads an existing model and scaler if available,
    or initializes new ones if not. It allows the model to be 
    incrementally trained with new data.
    """
    def __init__(self, model_path, scaler_path):
        """
        Initializes the IncrementalTrainer class.
        
        Args:
            model_path (str): Path to the model file.
            scaler_path (str): Path to the scaler file.
        """
        # Load existing model and scaler if they exist, otherwise initialize new ones
        if os.path.exists(model_path):
            self.model = joblib.load(model_path)
        else:
            self.model = SGDClassifier(max_iter=1000, tol=1e-3)

        if os.path.exists(scaler_path):
            self.scaler = joblib.load(scaler_path)
        else:
            self.scaler = StandardScaler()

    def fit(self, X, y):
        """
        Incrementally fits the model with new data.
        
        Args:
            X (numpy array): Feature matrix with shape (n_samples, n_features).
            y (numpy array): Target vector with shape (n_samples,).
        """
        self.scaler.partial_fit(X)
        X_scaled = self.scaler.transform(X)
        self.model.partial_fit(X_scaled, y, classes=np.unique(y))

    def save(self, model_path, scaler_path):
        """
        Saves the trained model and scaler to disk.
        
        Args:
            model_path (str): Path to save the model.
            scaler_path (str): Path to save the scaler.
        """
        joblib.dump(self.model, model_path)
        joblib.dump(self.scaler, scaler_path)

    def predict(self, X):
        """
        Predicts labels for new data.
        
        Args:
            X (numpy array): Feature matrix with shape (n_samples, n_features).
        
        Returns:
            numpy array: Predicted labels with shape (n_samples,).
        """
        X_scaled = self.scaler.transform(X)
        return self.model.predict(X_scaled)

def log(message):
    """
    Logs a message with a timestamp.
    
    Args:
        message (str): Message to log.
    """
    with open("training_log.txt", "a") as log_file:
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        log_file.write(f"{timestamp} - {message}\n")
    print(f"{timestamp} - {message}")

def write_to_csv(filename, data):
    """
    Writes data to a CSV file.
    
    Args:
        filename (str): Path to the CSV file.
        data (list): Data to write.
    """
    with open(filename, 'a', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(data)

def write_to_excel(filename, data, sheet_name="Sheet1"):
    """
    Writes data to an Excel file.
    
    Args:
        filename (str): Path to the Excel file.
        data (list): Data to write.
        sheet_name (str): Name of the sheet to write to.
    """
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename)
    
    wb = openpyxl.load_workbook(filename)
    ws = wb[sheet_name]

    for row in data:
        ws.append(row)

    wb.save(filename)

def write_to_json(filename, data):
    """
    Writes data to a JSON file.
    
    Args:
        filename (str): Path to the JSON file.
        data (dict): Data to write.
    """
    if os.path.exists(filename):
        with open(filename, 'r') as json_file:
            existing_data = json.load(json_file)
        existing_data.append(data)
    else:
        existing_data = [data]

    with open(filename, 'w') as json_file:
        json.dump(existing_data, json_file, indent=4)

def load_new_data(data_file='data/nmap_data.json'):
    """
    Loads new data from a JSON file for training.
    
    Args:
        data_file (str): Path to the JSON file.
    
    Returns:
        tuple: (X_new, y_vuln, y_threat, y_anomaly)
            X_new (numpy array): New feature matrix.
            y_vuln (numpy array): Vulnerability labels.
            y_threat (numpy array): Threat labels.
            y_anomaly (numpy array): Anomaly labels.
    """
    if os.path.exists(data_file):
        with open(data_file, 'r') as f:
            data = json.load(f)
        X_new = np.array([d['nmap_results'] for d in data])
        y_vuln = np.array([d['vulnerabilities'] for d in data])
        y_threat = np.array([d['threats'] for d in data])
        y_anomaly = np.array([d['anomalies'] for d in data])
        return X_new, y_vuln, y_threat, y_anomaly
    else:
        return None, None, None, None

def evaluate_model(model, scaler, X_test, y_test, prev_accuracy=None, accuracy_history=None, output_format="csv"):
    """
    Evaluates the model's performance and logs the results.
    
    Args:
        model: Trained model.
        scaler: Scaler used to normalize the data.
        X_test (numpy array): Test feature matrix.
        y_test (numpy array): Test target vector.
        prev_accuracy (float): Previous accuracy for comparison.
        accuracy_history (list): History of accuracy scores.
        output_format (str): Format for saving results ('csv', 'xlsx', 'json').
    
    Returns:
        float: Current accuracy score.
    """
    X_test_scaled = scaler.transform(X_test)
    y_pred = model.predict(X_test_scaled)
    accuracy = accuracy_score(y_test, y_pred)
    precision = precision_score(y_test, y_pred, average='weighted')
    recall = recall_score(y_test, y_pred, average='weighted')
    f1 = f1_score(y_test, y_pred, average='weighted')

    report = classification_report(y_test, y_pred, output_dict=True)
    cm = confusion_matrix(y_test, y_pred)

    # Plot confusion matrix
    plt.figure(figsize=(10, 7))
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues')
    plt.title('Confusion Matrix')
    plt.xlabel('Predicted')
    plt.ylabel('Actual')
    plt.show()

    # Print classification report with detailed explanations
    log(f"Model Evaluation Report:")
    log(f"Accuracy: {accuracy:.2f} - The proportion of true results (both true positives and true negatives) among the total number of cases examined.")
    log(f"Precision: {precision:.2f} - The ratio of correctly predicted positive observations to the total predicted positives.")
    log(f"Recall: {recall:.2f} - The ratio of correctly predicted positive observations to all observations in actual class.")
    log(f"F1-Score: {f1:.2f} - The weighted average of Precision and Recall.")
    log(f"Classification Report (detailed metrics for each class):\n{json.dumps(report, indent=2)}")

    for label, metrics in report.items():
        if isinstance(metrics, dict):
            log(f"Class {label}:")
            log(f"  Precision: {metrics['precision']:.2f} - The ratio of correctly predicted positive observations to the total predicted positives.")
            log(f"  Recall: {metrics['recall']:.2f} - The ratio of correctly predicted positive observations to all observations in actual class.")
            log(f"  F1-Score: {metrics['f1-score']:.2f} - The weighted average of Precision and Recall.")
            log(f"  Support: {metrics['support']} - The number of actual occurrences of the class in the provided dataset.")

    # Highlight significant changes with color coding
    if prev_accuracy is not None:
        change = accuracy - prev_accuracy
        change_percent = (change / prev_accuracy) * 100 if prev_accuracy != 0 else 0
        if change > 0:
            change_color = "\033[1;32m"  # Green for improvement
        elif change < 0:
            change_color = "\033[1;31m"  # Red for decline
        else:
            change_color = "\033[1;33m"  # Yellow for no change
        print(f"{change_color}Change in accuracy: {change:.2f} ({change_percent:.2f}%)\033[0m")

    # Update accuracy history
    if accuracy_history is not None:
        accuracy_history.append(accuracy)
        plt.figure(figsize=(10, 7))
        plt.plot(accuracy_history, marker='o')
        plt.title('Model Accuracy Over Time')
        plt.xlabel('Training Iterations')
        plt.ylabel('Accuracy')
        plt.grid(True)
        plt.show()

    # ROC Curve and AUC
    y_prob = model.decision_function(X_test_scaled)
    fpr, tpr, _ = roc_curve(y_test, y_prob)
    roc_auc = auc(fpr, tpr)

    plt.figure(figsize=(10, 7))
    plt.plot(fpr, tpr, color='darkorange', lw=2, label=f'ROC curve (area = {roc_auc:.2f})')
    plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
    plt.xlim([0.0, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel('False Positive Rate')
    plt.ylabel('True Positive Rate')
    plt.title('Receiver Operating Characteristic (ROC) Curve')
    plt.legend(loc="lower right")
    plt.grid(True)
    plt.show()

    # Write detailed results to selected format
    data = [
        time.strftime("%Y-%m-%d %H:%M:%S"),
        "Model Name",
        accuracy,
        report['0']['precision'],
        report['0']['recall'],
        report['0']['f1-score'],
        report['1']['precision'],
        report['1']['recall'],
        report['1']['f1-score'],
    ]

    if output_format == "csv":
        write_to_csv("model_evaluation_results.csv", data)
    elif output_format == "xlsx":
        write_to_excel("model_evaluation_results.xlsx", [data])
    elif output_format == "json":
        write_to_json("model_evaluation_results.json", {
            "timestamp": data[0],
            "model_name": data[1],
            "accuracy": data[2],
            "precision_class_0": data[3],
            "recall_class_0": data[4],
            "f1_score_class_0": data[5],
            "precision_class_1": data[6],
            "recall_class_1": data[7],
            "f1_score_class_1": data[8]
        })

    return accuracy

def incremental_training(output_format="csv"):
    """
    Performs incremental training and evaluation of models.
    
    Args:
        output_format (str): Format for saving results ('csv', 'xlsx', 'json').
    """
    X_new, y_vuln, y_threat, y_anomaly = load_new_data()

    if X_new is not None:
        accuracy_file = 'models/accuracy.json'
        if os.path.exists(accuracy_file):
            with open(accuracy_file, 'r') as f:
                accuracy_data = json.load(f)
            vuln_accuracy_history = accuracy_data.get('vulnerability_accuracy', [])
            threat_accuracy_history = accuracy_data.get('threat_accuracy', [])
            anomaly_accuracy_history = accuracy_data.get('anomaly_accuracy', [])
        else:
            vuln_accuracy_history = []
            threat_accuracy_history = []
            anomaly_accuracy_history = []

        # Create IncrementalTrainer instances with model and scaler paths
        trainer_vuln = IncrementalTrainer('models/vulnerability_model.pkl', 'models/vuln_scaler.pkl')
        trainer_threat = IncrementalTrainer('models/threat_model.pkl', 'models/threat_scaler.pkl')
        trainer_anomaly = IncrementalTrainer('models/anomaly_model.pkl', 'models/anomaly_scaler.pkl')

        # Train models incrementally with new data
        trainer_vuln.fit(X_new, y_vuln)
        trainer_threat.fit(X_new, y_threat)
        trainer_anomaly.fit(X_new, y_anomaly)

        # Save updated models and scalers
        trainer_vuln.save('models/vulnerability_model.pkl', 'models/vuln_scaler.pkl')
        trainer_threat.save('models/threat_model.pkl', 'models/threat_scaler.pkl')
        trainer_anomaly.save('models/anomaly_model.pkl', 'models/anomaly_scaler.pkl')

        # Placeholder test data for evaluation
        X_test = np.random.rand(10, 5)
        y_vuln_test = np.random.randint(0, 2, 10)
        y_threat_test = np.random.randint(0, 2, 10)
        y_anomaly_test = np.random.randint(0, 2, 10)

        # Get previous accuracies for comparison
        prev_vuln_accuracy = vuln_accuracy_history[-1] if vuln_accuracy_history else None
        prev_threat_accuracy = threat_accuracy_history[-1] if threat_accuracy_history else None
        prev_anomaly_accuracy = anomaly_accuracy_history[-1] if anomaly_accuracy_history else None

        # Evaluate models and get current accuracies
        current_vuln_accuracy = evaluate_model(trainer_vuln.model, trainer_vuln.scaler, X_test, y_vuln_test, prev_vuln_accuracy, vuln_accuracy_history, output_format)
        current_threat_accuracy = evaluate_model(trainer_threat.model, trainer_threat.scaler, X_test, y_threat_test, prev_threat_accuracy, threat_accuracy_history, output_format)
        current_anomaly_accuracy = evaluate_model(trainer_anomaly.model, trainer_anomaly.scaler, X_test, y_anomaly_test, prev_anomaly_accuracy, anomaly_accuracy_history, output_format)

        # Save updated accuracy histories
        with open(accuracy_file, 'w') as f:
            json.dump({
                'vulnerability_accuracy': vuln_accuracy_history,
                'threat_accuracy': threat_accuracy_history,
                'anomaly_accuracy': anomaly_accuracy_history
            }, f)

def main():
    """
    Main function to execute the incremental training and evaluation.
    """
    output_format = input("Enter the desired output format (csv, xlsx, json): ").strip().lower()
    if output_format not in {"csv", "xlsx", "json"}:
        print("Invalid format. Defaulting to CSV.")
        output_format = "csv"

    if output_format == "csv":
        write_to_csv("model_evaluation_results.csv", [
            "Timestamp", "Model Name", "Accuracy", "Precision (Class 0)", "Recall (Class 0)",
            "F1-Score (Class 0)", "Precision (Class 1)", "Recall (Class 1)", "F1-Score (Class 1)"
        ])
    elif output_format == "xlsx":
        write_to_excel("model_evaluation_results.xlsx", [[
            "Timestamp", "Model Name", "Accuracy", "Precision (Class 0)", "Recall (Class 0)",
            "F1-Score (Class 0)", "Precision (Class 1)", "Recall (Class 1)", "F1-Score (Class 1)"
        ]])

    incremental_training(output_format)

if __name__ == "__main__":
    main()
